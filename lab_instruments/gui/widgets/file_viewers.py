"""Read-only file viewers: image, CSV, PDF, DOCX, XLSX, PPTX, plain text."""

from __future__ import annotations

import contextlib
import csv
import hashlib
import os
import re
import shutil
import subprocess
import tempfile
from pathlib import Path


def _init_mupdf() -> None:
    """Suppress MuPDF C-level error output (e.g. structure tree warnings).

    Uses PyMuPDF's own API so suppression works across threads and doesn't
    redirect fd 2 globally.  Safe to call multiple times.
    """
    try:
        import fitz

        fitz.TOOLS.mupdf_display_errors(False)
        fitz.TOOLS.mupdf_warnings()  # clear any buffered warnings
    except Exception:
        pass


@contextlib.contextmanager
def _silence_mupdf():
    """Suppress MuPDF C-level stderr spam (e.g. structure tree warnings).

    Belt-and-suspenders: calls the PyMuPDF API *and* redirects fd 2 so that
    warnings printed before the API takes effect are also hidden.
    """
    _init_mupdf()
    devnull = os.open(os.devnull, os.O_WRONLY)
    saved = os.dup(2)
    os.dup2(devnull, 2)
    os.close(devnull)
    try:
        yield
    finally:
        os.dup2(saved, 2)
        os.close(saved)
        try:
            import fitz

            fitz.TOOLS.mupdf_warnings()  # clear any warnings accumulated during the call
        except Exception:
            pass


from PySide6.QtCore import QEvent, Qt, QThread, Signal
from PySide6.QtGui import QColor, QImage, QKeySequence, QPixmap, QShortcut, QTextCharFormat, QTextCursor
from PySide6.QtWidgets import (
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QPlainTextEdit,
    QPushButton,
    QScrollArea,
    QSizePolicy,
    QSpacerItem,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

from ..core.helpers import _mono

# Suppress MuPDF C-level error output as early as possible so that warnings
# like "No common ancestor in structure tree" never reach the terminal.
_init_mupdf()


class _FindBar(QWidget):
    """Slim search bar that embeds at the bottom of a viewer."""

    find_next = Signal(str)
    find_prev = Signal(str)

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        lay = QHBoxLayout(self)
        lay.setContentsMargins(4, 2, 4, 2)
        lay.setSpacing(4)

        self._input = QLineEdit()
        self._input.setPlaceholderText("Find…  (Esc to close)")
        self._input.setMaximumWidth(300)
        self._input.returnPressed.connect(lambda: self.find_next.emit(self._input.text()))
        self._input.textChanged.connect(lambda t: self.find_next.emit(t))
        lay.addWidget(self._input)

        prev_btn = QPushButton("↑")
        prev_btn.setFixedWidth(26)
        prev_btn.setToolTip("Previous match (Shift+Enter)")
        prev_btn.clicked.connect(lambda: self.find_prev.emit(self._input.text()))
        lay.addWidget(prev_btn)

        next_btn = QPushButton("↓")
        next_btn.setFixedWidth(26)
        next_btn.setToolTip("Next match (Enter)")
        next_btn.clicked.connect(lambda: self.find_next.emit(self._input.text()))
        lay.addWidget(next_btn)

        self._regex_btn = QPushButton(".*")
        self._regex_btn.setFixedWidth(28)
        self._regex_btn.setCheckable(True)
        self._regex_btn.setToolTip("Use regular expression")
        self._regex_btn.toggled.connect(lambda _: self.find_next.emit(self._input.text()))
        lay.addWidget(self._regex_btn)

        self._case_btn = QPushButton("Aa")
        self._case_btn.setFixedWidth(28)
        self._case_btn.setCheckable(True)
        self._case_btn.setToolTip("Match case")
        self._case_btn.toggled.connect(lambda _: self.find_next.emit(self._input.text()))
        lay.addWidget(self._case_btn)

        self._status = QLabel("")
        self._status.setStyleSheet("font-size: 10px; color: gray;")
        lay.addWidget(self._status)

        lay.addStretch()

        close_btn = QPushButton("✕")
        close_btn.setFixedWidth(26)
        close_btn.clicked.connect(self.hide)
        lay.addWidget(close_btn)

        self.hide()

    def open(self) -> None:
        self.show()
        self._input.setFocus()
        self._input.selectAll()

    def text(self) -> str:
        return self._input.text()

    def use_regex(self) -> bool:
        return self._regex_btn.isChecked()

    def case_sensitive(self) -> bool:
        return self._case_btn.isChecked()

    def set_status(self, text: str) -> None:
        self._status.setText(text)

    def keyPressEvent(self, event) -> None:  # noqa: N802
        if event.key() == Qt.Key.Key_Escape:
            self.hide()
        else:
            super().keyPressEvent(event)


class _TextFindReplaceBar(QWidget):
    """Find-and-replace bar for editable text viewers.

    Signals
    -------
    search_changed(pattern, use_regex, case_sensitive)
        Emitted whenever the search text or options change.
    find_next_requested / find_prev_requested
        Emitted when the user requests navigation.
    replace_one_requested(replacement)
        Emitted when the "Replace" button is clicked.
    replace_all_requested(replacement)
        Emitted when the "Replace All" button is clicked.
    closed
        Emitted when the bar is hidden (Esc or ✕).
    """

    search_changed = Signal(str, bool, bool)  # pattern, use_regex, case_sensitive
    find_next_requested = Signal()
    find_prev_requested = Signal()
    replace_one_requested = Signal(str)
    replace_all_requested = Signal(str)
    closed = Signal()

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        outer = QVBoxLayout(self)
        outer.setContentsMargins(4, 2, 4, 2)
        outer.setSpacing(2)

        # ── find row ─────────────────────────────────────────────────────────
        find_row = QHBoxLayout()
        find_row.setContentsMargins(0, 0, 0, 0)
        find_row.setSpacing(4)

        self._find_input = QLineEdit()
        self._find_input.setPlaceholderText("Find…  (Esc to close)")
        self._find_input.setMaximumWidth(300)
        find_row.addWidget(self._find_input)

        self._regex_btn = QPushButton(".*")
        self._regex_btn.setFixedWidth(32)
        self._regex_btn.setCheckable(True)
        self._regex_btn.setToolTip("Enable regex search")
        find_row.addWidget(self._regex_btn)

        self._case_btn = QPushButton("Aa")
        self._case_btn.setFixedWidth(32)
        self._case_btn.setCheckable(True)
        self._case_btn.setToolTip("Case-sensitive search")
        find_row.addWidget(self._case_btn)

        prev_btn = QPushButton("↑")
        prev_btn.setFixedWidth(26)
        prev_btn.setToolTip("Previous match")
        prev_btn.clicked.connect(self.find_prev_requested)
        find_row.addWidget(prev_btn)

        next_btn = QPushButton("↓")
        next_btn.setFixedWidth(26)
        next_btn.setToolTip("Next match")
        next_btn.clicked.connect(self.find_next_requested)
        find_row.addWidget(next_btn)

        self._status = QLabel("")
        self._status.setStyleSheet("font-size: 10px; color: gray;")
        find_row.addWidget(self._status)

        find_row.addStretch()

        close_btn = QPushButton("✕")
        close_btn.setFixedWidth(26)
        close_btn.clicked.connect(self._close)
        find_row.addWidget(close_btn)

        outer.addLayout(find_row)

        # ── replace row (hidden until Ctrl+H) ────────────────────────────────
        self._replace_row_widget = QWidget()
        repl_row = QHBoxLayout(self._replace_row_widget)
        repl_row.setContentsMargins(0, 0, 0, 0)
        repl_row.setSpacing(4)

        self._replace_input = QLineEdit()
        self._replace_input.setPlaceholderText("Replace with…")
        self._replace_input.setMaximumWidth(300)
        repl_row.addWidget(self._replace_input)

        repl_btn = QPushButton("Replace")
        repl_btn.clicked.connect(lambda: self.replace_one_requested.emit(self._replace_input.text()))
        repl_row.addWidget(repl_btn)

        repl_all_btn = QPushButton("Replace All")
        repl_all_btn.clicked.connect(lambda: self.replace_all_requested.emit(self._replace_input.text()))
        repl_row.addWidget(repl_all_btn)

        repl_row.addStretch()
        outer.addWidget(self._replace_row_widget)
        self._replace_row_widget.hide()

        # ── wire up signals ───────────────────────────────────────────────────
        self._find_input.textChanged.connect(self._emit_search_changed)
        self._find_input.returnPressed.connect(self.find_next_requested)
        self._regex_btn.toggled.connect(self._emit_search_changed)
        self._case_btn.toggled.connect(self._emit_search_changed)

        self.hide()

    # ── public API ────────────────────────────────────────────────────────────

    def open_find(self) -> None:
        """Show bar with focus on find field; replace row stays hidden."""
        self._replace_row_widget.hide()
        self.show()
        self._find_input.setFocus()
        self._find_input.selectAll()

    def open_replace(self) -> None:
        """Show bar with both find and replace rows."""
        self._replace_row_widget.show()
        self.show()
        self._find_input.setFocus()
        self._find_input.selectAll()

    def pattern(self) -> str:
        return self._find_input.text()

    def use_regex(self) -> bool:
        return self._regex_btn.isChecked()

    def case_sensitive(self) -> bool:
        return self._case_btn.isChecked()

    def set_status(self, text: str) -> None:
        self._status.setText(text)

    # ── internal helpers ─────────────────────────────────────────────────────

    def _emit_search_changed(self, *_) -> None:
        self.search_changed.emit(self.pattern(), self.use_regex(), self.case_sensitive())

    def _close(self) -> None:
        self.hide()
        self.closed.emit()

    def keyPressEvent(self, event) -> None:  # noqa: N802
        if event.key() == Qt.Key.Key_Escape:
            self._close()
        else:
            super().keyPressEvent(event)


class ImageViewer(QWidget):
    """Displays an image file (PNG, JPG, etc.) with basic zoom."""

    def __init__(self, file_path: str, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._path = file_path
        self._scale = 1.0

        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)

        # Toolbar
        tb = QHBoxLayout()
        tb.setContentsMargins(4, 4, 4, 4)
        for text, delta in [("Zoom In (+)", 0.25), ("Zoom Out (-)", -0.25), ("Fit", 0)]:
            b = QPushButton(text)
            b.clicked.connect(lambda _, d=delta: self._zoom(d))
            tb.addWidget(b)
        tb.addStretch()
        self._info = QLabel("")
        self._info.setStyleSheet("font-size: 10px;")
        tb.addWidget(self._info)
        lay.addLayout(tb)

        # Scroll area with image
        self._scroll = QScrollArea()
        self._scroll.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._label = QLabel()
        self._label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._scroll.setWidget(self._label)
        self._scroll.setWidgetResizable(False)
        lay.addWidget(self._scroll, 1)

        self._pixmap = QPixmap(file_path)
        if not self._pixmap.isNull():
            self._info.setText(
                f"{self._pixmap.width()}x{self._pixmap.height()}  {os.path.getsize(file_path) // 1024} KB"
            )
            self._label.setPixmap(self._pixmap)
            self._label.resize(self._pixmap.size())

        self.setFocusPolicy(Qt.FocusPolicy.StrongFocus)

    def keyPressEvent(self, event) -> None:  # noqa: N802
        key = event.key()
        if key in (Qt.Key.Key_Plus, Qt.Key.Key_Equal):
            self._zoom(0.25)
        elif key == Qt.Key.Key_Minus:
            self._zoom(-0.25)
        elif key in (Qt.Key.Key_F, Qt.Key.Key_0):
            self._zoom(0)
        else:
            super().keyPressEvent(event)

    def _zoom(self, delta: float) -> None:
        if delta == 0:
            # Fit to viewport
            vp = self._scroll.viewport().size()
            pw, ph = self._pixmap.width(), self._pixmap.height()
            if pw > 0 and ph > 0:
                self._scale = min(vp.width() / pw, vp.height() / ph)
        else:
            self._scale = max(0.1, min(5.0, self._scale + delta))
        scaled = self._pixmap.scaled(
            int(self._pixmap.width() * self._scale),
            int(self._pixmap.height() * self._scale),
            Qt.AspectRatioMode.KeepAspectRatio,
            Qt.TransformationMode.SmoothTransformation,
        )
        self._label.setPixmap(scaled)
        self._label.resize(scaled.size())


class CsvViewer(QWidget):
    """Displays a CSV file in a read-only table."""

    MAX_ROWS = 10000

    def __init__(self, file_path: str, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._path = file_path
        self._matches: list[tuple[int, int]] = []
        self._match_idx = 0

        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)

        self._table = QTableWidget()
        self._table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._table.setAlternatingRowColors(True)
        lay.addWidget(self._table, 1)

        self._info = QLabel("")
        self._info.setStyleSheet("font-size: 10px; padding: 2px 6px;")
        lay.addWidget(self._info)

        self._find_bar = _FindBar()
        self._find_bar.find_next.connect(self._find)
        self._find_bar.find_prev.connect(lambda t: self._find(t, backward=True))
        lay.addWidget(self._find_bar)

        _csv_find_sc = QShortcut(QKeySequence.StandardKey.Find, self)
        _csv_find_sc.setContext(Qt.ShortcutContext.WidgetWithChildrenShortcut)
        _csv_find_sc.activated.connect(self._find_bar.open)

        self._load()

    def _load(self) -> None:
        try:
            with open(self._path, newline="", encoding="utf-8") as f:
                reader = csv.reader(f)
                rows = []
                for i, row in enumerate(reader):
                    rows.append(row)
                    if i >= self.MAX_ROWS:
                        break
            if not rows:
                self._info.setText("Empty CSV")
                return
            # Use first row as header
            headers = rows[0]
            data = rows[1:]
            self._table.setColumnCount(len(headers))
            self._table.setHorizontalHeaderLabels(headers)
            self._table.setRowCount(len(data))
            for r, row in enumerate(data):
                for c, val in enumerate(row):
                    self._table.setItem(r, c, QTableWidgetItem(val))
            self._table.resizeColumnsToContents()
            self._info.setText(f"{len(data)} rows x {len(headers)} columns")
        except Exception as exc:
            self._info.setText(f"Error: {exc}")

    def _find(self, text: str, backward: bool = False) -> None:
        if not text:
            self._matches = []
            self._find_bar.set_status("")
            return
        use_regex = self._find_bar.use_regex()
        case_sensitive = self._find_bar.case_sensitive()
        matches = []
        for r in range(self._table.rowCount()):
            for c in range(self._table.columnCount()):
                item = self._table.item(r, c)
                if not item:
                    continue
                cell = item.text()
                if use_regex:
                    flags = 0 if case_sensitive else re.IGNORECASE
                    try:
                        if re.search(text, cell, flags):
                            matches.append((r, c))
                    except re.error:
                        self._find_bar.set_status("Regex error")
                        return
                else:
                    needle = text if case_sensitive else text.lower()
                    haystack = cell if case_sensitive else cell.lower()
                    if needle in haystack:
                        matches.append((r, c))
        self._matches = matches
        if not self._matches:
            self._find_bar.set_status("No results")
            return
        if backward:
            self._match_idx = (self._match_idx - 1) % len(self._matches)
        else:
            self._match_idx = (self._match_idx + 1) % len(self._matches) if self._matches else 0
        r, c = self._matches[self._match_idx]
        self._table.setCurrentCell(r, c)
        self._table.scrollToItem(self._table.item(r, c))
        self._find_bar.set_status(f"{self._match_idx + 1} of {len(self._matches)}")


class TextViewer(QWidget):
    """Editable plain text viewer with find+replace (including regex support)."""

    # Background colour for non-current matches
    _MATCH_BG = QColor("#ffe599")
    # Background colour for the current (active) match
    _CURRENT_BG = QColor("#f4a900")

    def __init__(self, file_path: str, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._path = file_path
        self._matches: list[tuple[int, int]] = []  # (start, end) char offsets
        self._match_idx: int = -1

        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)

        self._editor = QPlainTextEdit()
        # Editable — no setReadOnly
        self._editor.setFont(_mono())
        self._editor.setLineWrapMode(QPlainTextEdit.LineWrapMode.NoWrap)
        lay.addWidget(self._editor, 1)

        self._find_bar = _TextFindReplaceBar()
        self._find_bar.search_changed.connect(self._on_search_changed)
        self._find_bar.find_next_requested.connect(self._find_next)
        self._find_bar.find_prev_requested.connect(self._find_prev)
        self._find_bar.replace_one_requested.connect(self._replace_one)
        self._find_bar.replace_all_requested.connect(self._replace_all)
        self._find_bar.closed.connect(self._on_bar_closed)
        lay.addWidget(self._find_bar)

        # Keyboard shortcuts
        _find_sc = QShortcut(QKeySequence.StandardKey.Find, self)
        _find_sc.setContext(Qt.ShortcutContext.WidgetWithChildrenShortcut)
        _find_sc.activated.connect(self._find_bar.open_find)
        _repl_sc = QShortcut(QKeySequence("Ctrl+H"), self)
        _repl_sc.setContext(Qt.ShortcutContext.WidgetWithChildrenShortcut)
        _repl_sc.activated.connect(self._find_bar.open_replace)
        _save_sc = QShortcut(QKeySequence.StandardKey.Save, self)
        _save_sc.setContext(Qt.ShortcutContext.WidgetWithChildrenShortcut)
        _save_sc.activated.connect(self._save)

        try:
            with open(file_path, encoding="utf-8", errors="replace") as f:
                self._editor.setPlainText(f.read())
        except Exception as exc:
            self._editor.setPlainText(f"Error reading file: {exc}")

    # ── file save ─────────────────────────────────────────────────────────────

    def _save(self) -> None:
        try:
            with open(self._path, "w", encoding="utf-8") as f:
                f.write(self._editor.toPlainText())
        except Exception as exc:
            self._find_bar.set_status(f"Save error: {exc}")

    # ── search changed (re-scan matches) ─────────────────────────────────────

    def _on_search_changed(self, pattern: str, use_regex: bool, case_sensitive: bool) -> None:
        self._update_matches(pattern, use_regex, case_sensitive)

    def _update_matches(
        self, pattern: str | None = None, use_regex: bool | None = None, case_sensitive: bool | None = None
    ) -> None:
        """Rebuild match list from current bar settings and highlight all."""
        if pattern is None:
            pattern = self._find_bar.pattern()
        if use_regex is None:
            use_regex = self._find_bar.use_regex()
        if case_sensitive is None:
            case_sensitive = self._find_bar.case_sensitive()

        self._matches = []
        self._match_idx = -1

        if not pattern:
            self._highlight_all()
            self._find_bar.set_status("")
            return

        text = self._editor.toPlainText()

        if use_regex:
            flags = 0 if case_sensitive else re.IGNORECASE
            try:
                compiled = re.compile(pattern, flags)
            except re.error as exc:
                self._find_bar.set_status(f"Regex error: {exc}")
                self._highlight_all()
                return
            self._matches = [(m.start(), m.end()) for m in compiled.finditer(text)]
        else:
            search = pattern if case_sensitive else pattern.lower()
            haystack = text if case_sensitive else text.lower()
            start = 0
            while True:
                idx = haystack.find(search, start)
                if idx == -1:
                    break
                self._matches.append((idx, idx + len(search)))
                start = idx + max(1, len(search))

        if self._matches:
            self._match_idx = 0
            self._go_to_match(0)
        else:
            self._highlight_all()
            self._find_bar.set_status("No results")

    # ── navigation ────────────────────────────────────────────────────────────

    def _find_next(self) -> None:
        if not self._matches:
            self._update_matches()
            return
        self._match_idx = (self._match_idx + 1) % len(self._matches)
        self._go_to_match(self._match_idx)

    def _find_prev(self) -> None:
        if not self._matches:
            self._update_matches()
            return
        self._match_idx = (self._match_idx - 1) % len(self._matches)
        self._go_to_match(self._match_idx)

    def _go_to_match(self, idx: int) -> None:
        if not self._matches:
            return
        start, end = self._matches[idx]
        doc = self._editor.document()
        cursor = QTextCursor(doc)
        cursor.setPosition(start)
        cursor.setPosition(end, QTextCursor.MoveMode.KeepAnchor)
        self._editor.setTextCursor(cursor)
        self._editor.ensureCursorVisible()
        self._highlight_all()
        self._find_bar.set_status(f"{idx + 1} of {len(self._matches)}")

    # ── highlighting ─────────────────────────────────────────────────────────

    def _highlight_all(self) -> None:
        """Apply extra selections: pale yellow for all matches, amber for current."""
        selections: list[QTextEdit.ExtraSelection] = []
        doc = self._editor.document()

        for i, (start, end) in enumerate(self._matches):
            sel = QTextEdit.ExtraSelection()
            fmt = QTextCharFormat()
            if i == self._match_idx:
                fmt.setBackground(self._CURRENT_BG)
            else:
                fmt.setBackground(self._MATCH_BG)
            sel.format = fmt
            cur = QTextCursor(doc)
            cur.setPosition(start)
            cur.setPosition(end, QTextCursor.MoveMode.KeepAnchor)
            sel.cursor = cur
            selections.append(sel)

        self._editor.setExtraSelections(selections)

    # ── replace ───────────────────────────────────────────────────────────────

    def _replace_one(self, replacement: str) -> None:
        """Replace the current match and advance to the next."""
        if not self._matches or self._match_idx < 0:
            self._update_matches()
            return
        start, end = self._matches[self._match_idx]
        pattern = self._find_bar.pattern()
        use_regex = self._find_bar.use_regex()
        case_sensitive = self._find_bar.case_sensitive()

        if use_regex:
            flags = 0 if case_sensitive else re.IGNORECASE
            try:
                compiled = re.compile(pattern, flags)
            except re.error:
                return
            text = self._editor.toPlainText()
            matched_text = text[start:end]
            try:
                replacement_text = compiled.sub(replacement, matched_text, count=1)
            except re.error:
                return
        else:
            replacement_text = replacement

        doc = self._editor.document()
        cursor = QTextCursor(doc)
        cursor.setPosition(start)
        cursor.setPosition(end, QTextCursor.MoveMode.KeepAnchor)
        cursor.insertText(replacement_text)

        # Re-scan and advance
        self._update_matches()
        if self._matches:
            self._find_next()

    def _replace_all(self, replacement: str) -> None:
        """Replace all matches in the document."""
        pattern = self._find_bar.pattern()
        use_regex = self._find_bar.use_regex()
        case_sensitive = self._find_bar.case_sensitive()

        if not pattern:
            return

        text = self._editor.toPlainText()

        if use_regex:
            flags = 0 if case_sensitive else re.IGNORECASE
            try:
                compiled = re.compile(pattern, flags)
                count = len(compiled.findall(text))
                new_text = compiled.sub(replacement, text)
            except re.error as exc:
                self._find_bar.set_status(f"Regex error: {exc}")
                return
        else:
            search = pattern if case_sensitive else pattern.lower()
            haystack = text if case_sensitive else text.lower()
            count = haystack.count(search)
            if case_sensitive:
                new_text = text.replace(pattern, replacement)
            else:
                # Case-insensitive replace: rebuild via regex literal
                new_text = re.compile(re.escape(pattern), re.IGNORECASE).sub(replacement, text)

        self._editor.setPlainText(new_text)
        self._matches = []
        self._match_idx = -1
        self._highlight_all()
        self._find_bar.set_status(f"Replaced {count} occurrence{'s' if count != 1 else ''}")

    # ── bar closed ────────────────────────────────────────────────────────────

    def _on_bar_closed(self) -> None:
        self._matches = []
        self._match_idx = -1
        self._editor.setExtraSelections([])
        self._editor.setFocus()


class PdfViewer(QWidget):
    """Renders PDF pages as images using PyMuPDF."""

    _ZOOM_MIN = 0.05
    _ZOOM_MAX = 20.0
    _ZOOM_STEP = 1.25

    def __init__(self, file_path: str, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._path = file_path
        self._page = 0
        self._page_count = 0
        self._zoom_scale = 2.0
        self._zoom_fit: str | None = "width"
        self.setFocusPolicy(Qt.FocusPolicy.StrongFocus)

        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)

        # Toolbar
        tb = QHBoxLayout()
        tb.setContentsMargins(4, 4, 4, 4)
        self._prev_btn = QPushButton("< Prev")
        self._prev_btn.clicked.connect(self._prev)
        tb.addWidget(self._prev_btn)
        self._next_btn = QPushButton("Next >")
        self._next_btn.clicked.connect(self._next)
        tb.addWidget(self._next_btn)
        tb.addStretch()

        fit_w_btn = QPushButton("Fit W")
        fit_w_btn.setFixedWidth(46)
        fit_w_btn.setToolTip("Fit width (W)")
        fit_w_btn.clicked.connect(self._fit_width)
        tb.addWidget(fit_w_btn)

        fit_h_btn = QPushButton("Fit H")
        fit_h_btn.setFixedWidth(46)
        fit_h_btn.setToolTip("Fit height (H)")
        fit_h_btn.clicked.connect(self._fit_height)
        tb.addWidget(fit_h_btn)

        zoom_out_btn = QPushButton("−")
        zoom_out_btn.setFixedWidth(26)
        zoom_out_btn.clicked.connect(lambda: self._zoom_by(1 / self._ZOOM_STEP))
        tb.addWidget(zoom_out_btn)

        self._zoom_label = QLabel("")
        self._zoom_label.setFixedWidth(54)
        self._zoom_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._zoom_label.setStyleSheet("font-size: 10px;")
        tb.addWidget(self._zoom_label)

        zoom_in_btn = QPushButton("+")
        zoom_in_btn.setFixedWidth(26)
        zoom_in_btn.clicked.connect(lambda: self._zoom_by(self._ZOOM_STEP))
        tb.addWidget(zoom_in_btn)

        self._info = QLabel("")
        self._info.setStyleSheet("font-size: 10px; margin-left: 8px;")
        tb.addWidget(self._info)
        lay.addLayout(tb)

        self._scroll = QScrollArea()
        self._scroll.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._scroll.setWidgetResizable(False)
        self._label = QLabel()
        self._label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._scroll.setWidget(self._label)
        self._scroll.installEventFilter(self)
        self._scroll.viewport().installEventFilter(self)
        lay.addWidget(self._scroll, 1)

        try:
            import fitz

            with _silence_mupdf():
                self._doc = fitz.open(file_path)
            self._page_count = len(self._doc)
            self._render()
        except ImportError:
            self._label.setText("Install pymupdf to view PDFs:\npip install pymupdf")
        except Exception as exc:
            self._label.setText(f"Error: {exc}")

    def _render(self) -> None:
        import fitz  # noqa: F811

        page = self._doc[self._page]
        mat = fitz.Matrix(self._zoom_scale, self._zoom_scale)
        with _silence_mupdf():
            pix = page.get_pixmap(matrix=mat)
        img = QImage(pix.samples, pix.width, pix.height, pix.stride, QImage.Format.Format_RGB888)
        pixmap = QPixmap.fromImage(img)
        self._label.setPixmap(pixmap)
        self._label.resize(pixmap.size())
        self._zoom_label.setText(f"{int(self._zoom_scale * 100)}%")
        self._info.setText(f"Page {self._page + 1} / {self._page_count}")
        self._prev_btn.setEnabled(self._page > 0)
        self._next_btn.setEnabled(self._page < self._page_count - 1)

    def _zoom_by(self, factor: float) -> None:
        self._zoom_scale = max(self._ZOOM_MIN, min(self._ZOOM_MAX, self._zoom_scale * factor))
        self._zoom_fit = None
        self._render()

    def _prev(self) -> None:
        if self._page > 0:
            self._page -= 1
            self._render()

    def _next(self) -> None:
        if self._page < self._page_count - 1:
            self._page += 1
            self._render()

    # _zoom_fit values: None = manual, 'width' = fit width, 'height' = fit height
    def _fit_width(self) -> None:
        if not self._page_count:
            return
        vp_w = self._scroll.viewport().width() - 4
        page = self._doc[self._page]
        if vp_w > 0 and page.rect.width > 0:
            self._zoom_scale = vp_w / page.rect.width
            self._zoom_fit = "width"
            self._render()

    def _fit_height(self) -> None:
        if not self._page_count:
            return
        vp_w = self._scroll.viewport().width() - 4
        vp_h = self._scroll.viewport().height() - 4
        page = self._doc[self._page]
        if vp_w > 0 and vp_h > 0 and page.rect.width > 0 and page.rect.height > 0:
            self._zoom_scale = min(vp_w / page.rect.width, vp_h / page.rect.height)
            self._zoom_fit = "height"
            self._render()

    def resizeEvent(self, event) -> None:  # noqa: N802
        if self._zoom_fit == "width" and self._page_count:
            self._fit_width()
        elif self._zoom_fit == "height" and self._page_count:
            self._fit_height()
        super().resizeEvent(event)

    def eventFilter(self, obj, event) -> bool:  # noqa: N802
        if obj is self._scroll.viewport() and event.type() == QEvent.Type.Wheel:
            if event.modifiers() & Qt.KeyboardModifier.ControlModifier:
                factor = self._ZOOM_STEP if event.angleDelta().y() > 0 else 1 / self._ZOOM_STEP
                self._zoom_by(factor)
                return True
        if event.type() == QEvent.Type.KeyPress:
            key = event.key()
            if key == Qt.Key.Key_Left:
                self._prev()
                return True
            if key == Qt.Key.Key_Right:
                self._next()
                return True
            if key in (Qt.Key.Key_Plus, Qt.Key.Key_Equal):
                self._zoom_by(self._ZOOM_STEP)
                return True
            if key == Qt.Key.Key_Minus:
                self._zoom_by(1 / self._ZOOM_STEP)
                return True
            if key == Qt.Key.Key_W:
                self._fit_width()
                return True
            if key == Qt.Key.Key_H:
                self._fit_height()
                return True
        return super().eventFilter(obj, event)

    def keyPressEvent(self, event) -> None:  # noqa: N802
        key = event.key()
        if key == Qt.Key.Key_Left:
            self._prev()
        elif key == Qt.Key.Key_Right:
            self._next()
        elif key in (Qt.Key.Key_Plus, Qt.Key.Key_Equal):
            self._zoom_by(self._ZOOM_STEP)
        elif key == Qt.Key.Key_Minus:
            self._zoom_by(1 / self._ZOOM_STEP)
        elif key == Qt.Key.Key_W:
            self._fit_width()
        elif key == Qt.Key.Key_H:
            self._fit_height()
        else:
            super().keyPressEvent(event)


# ---------------------------------------------------------------------------
# Shared helpers for Office document viewers (DOCX, XLSX, PPTX)
# ---------------------------------------------------------------------------


def _find_soffice() -> str | None:
    """Return the path to a working LibreOffice ``soffice`` binary, or None."""
    for name in ("soffice", "libreoffice"):
        p = shutil.which(name)
        if p:
            return p
    for fixed in (
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
        "/usr/bin/soffice",
        "/usr/local/bin/soffice",
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    ):
        if os.path.isfile(fixed):
            return fixed
    return None


def _office_cache_dir(file_path: str) -> Path:
    """Return a per-file cache directory for converted pages.

    The cache is keyed on path + mtime so edits to the source file
    automatically invalidate stale PNGs.
    """
    mtime = int(os.path.getmtime(file_path))
    key = f"{file_path}\0{mtime}"
    digest = hashlib.md5(key.encode()).hexdigest()[:16]  # noqa: S324
    d = Path(tempfile.gettempdir()) / "office_viewer_cache" / digest
    d.mkdir(parents=True, exist_ok=True)
    return d


def _soffice_user_profile() -> str:
    """Return a unique LibreOffice user-profile URI to avoid lock conflicts.

    soffice hangs when another instance is already running and they share
    the default profile.  Using a disposable profile per-process avoids this.
    """
    d = Path(tempfile.gettempdir()) / "office_viewer_profile" / str(os.getpid())
    d.mkdir(parents=True, exist_ok=True)
    return d.as_uri()


def _soffice_cmd(soffice_bin: str, cache_dir: Path, file_path: str) -> list[str]:
    """Build the soffice conversion command with a unique user profile."""
    return [
        soffice_bin,
        "--headless",
        "--nolockcheck",
        f"-env:UserInstallation={_soffice_user_profile()}",
        "--convert-to",
        "pdf",
        "--outdir",
        str(cache_dir),
        file_path,
    ]


class _OfficeConversionWorker(QThread):
    """Background thread: soffice -> PDF -> PyMuPDF -> QImage list."""

    pages_ready = Signal(list)
    error = Signal(str)
    progress = Signal(str)

    def __init__(self, file_path: str, soffice_bin: str, cache_dir: Path) -> None:
        super().__init__()
        self._file_path = file_path
        self._soffice = soffice_bin
        self._cache_dir = cache_dir

    def run(self) -> None:  # noqa: C901
        try:
            # 1. Cache hit?
            cached = sorted(self._cache_dir.glob("page_*.png"))
            if cached:
                images = []
                for p in cached:
                    img = QImage(str(p))
                    if not img.isNull():
                        images.append(img)
                if images:
                    self.pages_ready.emit(images)
                    return

            # 2. Convert to PDF via LibreOffice
            self.progress.emit("Converting with LibreOffice...")
            result = subprocess.run(
                _soffice_cmd(self._soffice, self._cache_dir, self._file_path),
                capture_output=True,
                timeout=180,
            )
            if result.returncode != 0:
                stderr = result.stderr.decode(errors="replace")[:400]
                self.error.emit(f"LibreOffice failed (exit {result.returncode}):\n{stderr}")
                return

            # 3. Find the PDF
            pdfs = list(self._cache_dir.glob("*.pdf"))
            if not pdfs:
                self.error.emit("LibreOffice did not produce a PDF file.")
                return
            pdf_path = pdfs[0]

            # 4. Render with PyMuPDF
            try:
                import fitz
            except ImportError:
                self.error.emit("pymupdf is not installed. Run: pip install pymupdf")
                return

            with _silence_mupdf():
                doc = fitz.open(str(pdf_path))
            images: list[QImage] = []
            for i, page in enumerate(doc):
                self.progress.emit(f"Rendering page {i + 1} of {len(doc)}...")
                mat = fitz.Matrix(2.0, 2.0)  # 144 DPI
                with _silence_mupdf():
                    pix = page.get_pixmap(matrix=mat, alpha=False)
                img = QImage(
                    pix.samples,
                    pix.width,
                    pix.height,
                    pix.stride,
                    QImage.Format.Format_RGB888,
                ).copy()
                img.save(str(self._cache_dir / f"page_{i:04d}.png"))
                images.append(img)
            doc.close()

            self.pages_ready.emit(images)
        except Exception as exc:
            self.error.emit(str(exc))


class _OfficeViewerBase(QWidget):
    """Base viewer for Office documents rendered via LibreOffice + PyMuPDF."""

    _ZOOM_MIN = 5.0
    _ZOOM_MAX = 2000.0
    _ZOOM_STEP = 1.25

    def __init__(self, file_path: str, page_label: str = "Page", parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._path = file_path
        self._page_label = page_label
        self._pages: list[QImage] = []
        self._current = 0
        self._worker: _OfficeConversionWorker | None = None
        self._zoom_pct: float = 100.0
        self._zoom_fit: str | None = "width"

        self.setFocusPolicy(Qt.FocusPolicy.StrongFocus)

        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)

        # -- toolbar --
        tb = QHBoxLayout()
        tb.setContentsMargins(4, 4, 4, 4)

        self._prev_btn = QPushButton("\u2039")
        self._prev_btn.setFixedWidth(32)
        self._prev_btn.setEnabled(False)
        self._prev_btn.clicked.connect(self._prev)
        tb.addWidget(self._prev_btn)

        self._page_info = QLabel("")
        self._page_info.setMinimumWidth(120)
        self._page_info.setAlignment(Qt.AlignmentFlag.AlignCenter)
        tb.addWidget(self._page_info)

        self._next_btn = QPushButton("\u203a")
        self._next_btn.setFixedWidth(32)
        self._next_btn.setEnabled(False)
        self._next_btn.clicked.connect(self._next)
        tb.addWidget(self._next_btn)

        tb.addItem(QSpacerItem(0, 0, QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum))

        fit_w_btn = QPushButton("Fit W")
        fit_w_btn.setFixedWidth(46)
        fit_w_btn.setToolTip("Fit width (W)")
        fit_w_btn.clicked.connect(self._fit_width)
        tb.addWidget(fit_w_btn)

        fit_h_btn = QPushButton("Fit H")
        fit_h_btn.setFixedWidth(46)
        fit_h_btn.setToolTip("Fit height (H)")
        fit_h_btn.clicked.connect(self._fit_height)
        tb.addWidget(fit_h_btn)

        zoom_out_btn = QPushButton("−")
        zoom_out_btn.setFixedWidth(26)
        zoom_out_btn.clicked.connect(lambda: self._zoom_by(1 / self._ZOOM_STEP))
        tb.addWidget(zoom_out_btn)

        self._zoom_label = QLabel("100%")
        self._zoom_label.setFixedWidth(54)
        self._zoom_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._zoom_label.setStyleSheet("font-size: 10px;")
        tb.addWidget(self._zoom_label)

        zoom_in_btn = QPushButton("+")
        zoom_in_btn.setFixedWidth(26)
        zoom_in_btn.clicked.connect(lambda: self._zoom_by(self._ZOOM_STEP))
        tb.addWidget(zoom_in_btn)

        lay.addLayout(tb)

        # -- scroll area --
        self._scroll = QScrollArea()
        self._scroll.setWidgetResizable(False)
        self._scroll.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._scroll.installEventFilter(self)
        self._scroll.viewport().installEventFilter(self)

        self._label = QLabel()
        self._label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._scroll.setWidget(self._label)
        lay.addWidget(self._scroll, 1)

        # -- kick off --
        soffice = _find_soffice()
        if soffice is None:
            self._label.setText(
                "LibreOffice is required to render Office documents.\n\n"
                "Install it from https://www.libreoffice.org -- it's free.\n"
                "Restart the app after installing."
            )
            self._label.setWordWrap(True)
            return

        self._label.setText("Converting... this may take a moment on first open.")
        cache = _office_cache_dir(file_path)
        self._worker = _OfficeConversionWorker(file_path, soffice, cache)
        self._worker.pages_ready.connect(self._on_pages_ready)
        self._worker.error.connect(self._on_error)
        self._worker.progress.connect(self._on_progress)
        self._worker.start()

    # -- slots --

    def _on_pages_ready(self, images: list[QImage]) -> None:
        self._pages = images
        self._current = 0
        self._go_to(0)

    def _on_error(self, msg: str) -> None:
        self._label.setText(msg)
        self._label.setWordWrap(True)
        self._label.setStyleSheet("color: #c0392b;")

    def _on_progress(self, msg: str) -> None:
        self._label.setText(msg)

    # -- navigation --

    def _go_to(self, index: int) -> None:
        if not self._pages:
            return
        self._current = max(0, min(index, len(self._pages) - 1))
        self._apply_zoom()
        total = len(self._pages)
        self._page_info.setText(f"{self._page_label} {self._current + 1} of {total}")
        self._prev_btn.setEnabled(self._current > 0)
        self._next_btn.setEnabled(self._current < total - 1)

    # -- zoom --

    def _fit_width(self) -> None:
        if not self._pages:
            return
        vp_w = self._scroll.viewport().width() - 4
        src = self._pages[self._current]
        if vp_w > 0 and src.width() > 0:
            self._zoom_pct = vp_w / src.width() * 100
            self._zoom_fit = "width"
            self._apply_zoom()

    def _fit_height(self) -> None:
        if not self._pages:
            return
        vp_w = self._scroll.viewport().width() - 4
        vp_h = self._scroll.viewport().height() - 4
        src = self._pages[self._current]
        if vp_w > 0 and vp_h > 0 and src.width() > 0 and src.height() > 0:
            self._zoom_pct = min(vp_w / src.width(), vp_h / src.height()) * 100
            self._zoom_fit = "height"
            self._apply_zoom()

    def _zoom_by(self, factor: float) -> None:
        self._zoom_pct = max(self._ZOOM_MIN, min(self._ZOOM_MAX, self._zoom_pct * factor))
        self._zoom_fit = None
        self._apply_zoom()

    def _apply_zoom(self) -> None:
        if not self._pages:
            return
        src = self._pages[self._current]
        if self._zoom_fit == "width":
            vp_w = self._scroll.viewport().width() - 4
            if vp_w > 0 and src.width() > 0:
                self._zoom_pct = vp_w / src.width() * 100
        elif self._zoom_fit == "height":
            vp_w = self._scroll.viewport().width() - 4
            vp_h = self._scroll.viewport().height() - 4
            if vp_w > 0 and vp_h > 0 and src.width() > 0 and src.height() > 0:
                self._zoom_pct = min(vp_w / src.width(), vp_h / src.height()) * 100
        target_w = max(1, int(src.width() * self._zoom_pct / 100))
        pix = QPixmap.fromImage(src).scaledToWidth(target_w, Qt.TransformationMode.SmoothTransformation)
        self._label.setPixmap(pix)
        self._label.resize(pix.size())
        self._zoom_label.setText(f"{int(self._zoom_pct)}%")

    def _prev(self) -> None:
        if self._current > 0:
            self._go_to(self._current - 1)

    def _next(self) -> None:
        if self._pages and self._current < len(self._pages) - 1:
            self._go_to(self._current + 1)

    # -- events --

    def resizeEvent(self, event) -> None:  # noqa: N802
        if self._zoom_fit:
            self._apply_zoom()
        super().resizeEvent(event)

    def eventFilter(self, obj, event) -> bool:  # noqa: N802
        if obj is self._scroll.viewport() and event.type() == QEvent.Type.Wheel:
            if event.modifiers() & Qt.KeyboardModifier.ControlModifier:
                factor = self._ZOOM_STEP if event.angleDelta().y() > 0 else 1 / self._ZOOM_STEP
                self._zoom_by(factor)
                return True
        if event.type() == QEvent.Type.KeyPress:
            key = event.key()
            if key == Qt.Key.Key_Left:
                self._prev()
                return True
            if key == Qt.Key.Key_Right:
                self._next()
                return True
            if key == Qt.Key.Key_Home:
                self._go_to(0)
                return True
            if key == Qt.Key.Key_End:
                self._go_to(len(self._pages) - 1)
                return True
            if key in (Qt.Key.Key_Plus, Qt.Key.Key_Equal):
                self._zoom_by(self._ZOOM_STEP)
                return True
            if key == Qt.Key.Key_Minus:
                self._zoom_by(1 / self._ZOOM_STEP)
                return True
            if key == Qt.Key.Key_W:
                self._fit_width()
                return True
            if key == Qt.Key.Key_H:
                self._fit_height()
                return True
        return super().eventFilter(obj, event)

    def keyPressEvent(self, event) -> None:  # noqa: N802
        key = event.key()
        if key == Qt.Key.Key_Left:
            self._prev()
        elif key == Qt.Key.Key_Right:
            self._next()
        elif key == Qt.Key.Key_Home:
            self._go_to(0)
        elif key == Qt.Key.Key_End:
            self._go_to(len(self._pages) - 1)
        elif key in (Qt.Key.Key_Plus, Qt.Key.Key_Equal):
            self._zoom_by(self._ZOOM_STEP)
        elif key == Qt.Key.Key_Minus:
            self._zoom_by(1 / self._ZOOM_STEP)
        elif key == Qt.Key.Key_W:
            self._fit_width()
        elif key == Qt.Key.Key_H:
            self._fit_height()
        else:
            super().keyPressEvent(event)

    def closeEvent(self, event) -> None:  # noqa: N802
        if self._worker is not None and self._worker.isRunning():
            self._worker.quit()
            self._worker.wait(3000)
        super().closeEvent(event)


# ---------------------------------------------------------------------------
# Public viewer classes (thin wrappers preserving original constructor API)
# ---------------------------------------------------------------------------


class DocxViewer(_OfficeViewerBase):
    """Displays .docx documents via LibreOffice PDF conversion."""

    def __init__(self, file_path: str, parent: QWidget | None = None) -> None:
        super().__init__(file_path, page_label="Page", parent=parent)


class XlsxViewer(QWidget):
    """Displays .xlsx spreadsheets in a read-only table with sheet tabs."""

    def __init__(self, file_path: str, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._matches: list[tuple[int, int]] = []
        self._match_idx = 0
        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)

        self._table = QTableWidget()
        self._table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._table.setAlternatingRowColors(True)
        lay.addWidget(self._table, 1)

        tb = QHBoxLayout()
        tb.setContentsMargins(4, 2, 4, 2)
        self._sheet_label = QLabel("")
        self._sheet_label.setStyleSheet("font-size: 10px;")
        tb.addWidget(self._sheet_label)
        tb.addStretch()
        self._info = QLabel("")
        self._info.setStyleSheet("font-size: 10px;")
        tb.addWidget(self._info)
        lay.addLayout(tb)

        self._find_bar = _FindBar()
        self._find_bar.find_next.connect(self._find)
        self._find_bar.find_prev.connect(lambda t: self._find(t, backward=True))
        lay.addWidget(self._find_bar)

        _xlsx_find_sc = QShortcut(QKeySequence.StandardKey.Find, self)
        _xlsx_find_sc.setContext(Qt.ShortcutContext.WidgetWithChildrenShortcut)
        _xlsx_find_sc.activated.connect(self._find_bar.open)

        try:
            from openpyxl import load_workbook

            self._wb = load_workbook(file_path, read_only=True, data_only=True)
            self._sheets = self._wb.sheetnames
            if self._sheets:
                self._load_sheet(self._sheets[0])
                self._sheet_label.setText(f"Sheets: {', '.join(self._sheets)}")
        except ImportError:
            self._info.setText("Install openpyxl: pip install openpyxl")
        except Exception as exc:
            self._info.setText(f"Error: {exc}")

    def _load_sheet(self, name: str) -> None:
        ws = self._wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            self._info.setText("Empty sheet")
            return
        ncols = max(len(r) for r in rows)
        self._table.setColumnCount(ncols)
        self._table.setRowCount(len(rows))
        for r, row in enumerate(rows):
            for c, val in enumerate(row or []):
                self._table.setItem(r, c, QTableWidgetItem(str(val) if val is not None else ""))
        self._table.resizeColumnsToContents()
        self._info.setText(f"{len(rows)} rows x {ncols} columns - {name}")

    def _find(self, text: str, backward: bool = False) -> None:
        if not text:
            self._matches = []
            self._find_bar.set_status("")
            return
        use_regex = self._find_bar.use_regex()
        case_sensitive = self._find_bar.case_sensitive()
        matches = []
        for r in range(self._table.rowCount()):
            for c in range(self._table.columnCount()):
                item = self._table.item(r, c)
                if not item:
                    continue
                cell = item.text()
                if use_regex:
                    flags = 0 if case_sensitive else re.IGNORECASE
                    try:
                        if re.search(text, cell, flags):
                            matches.append((r, c))
                    except re.error:
                        self._find_bar.set_status("Regex error")
                        return
                else:
                    needle = text if case_sensitive else text.lower()
                    haystack = cell if case_sensitive else cell.lower()
                    if needle in haystack:
                        matches.append((r, c))
        self._matches = matches
        if not self._matches:
            self._find_bar.set_status("No results")
            return
        if backward:
            self._match_idx = (self._match_idx - 1) % len(self._matches)
        else:
            self._match_idx = (self._match_idx + 1) % len(self._matches) if self._matches else 0
        r, c = self._matches[self._match_idx]
        self._table.setCurrentCell(r, c)
        self._table.scrollToItem(self._table.item(r, c))
        self._find_bar.set_status(f"{self._match_idx + 1} of {len(self._matches)}")


class PptxViewer(_OfficeViewerBase):
    """Renders .pptx slides via LibreOffice PDF conversion."""

    def __init__(self, file_path: str, parent: QWidget | None = None) -> None:
        super().__init__(file_path, page_label="Slide", parent=parent)


# ---------------------------------------------------------------------------
# Background pre-conversion of all Office files in a workspace
# ---------------------------------------------------------------------------

_OFFICE_EXTS = {".docx", ".doc", ".pptx"}


class _PreconvertWorker(QThread):
    """Walks a folder tree, converts every Office file to cached PNGs."""

    finished_all = Signal()

    def __init__(self, folder: str, soffice_bin: str) -> None:
        super().__init__()
        self._folder = folder
        self._soffice = soffice_bin

    def run(self) -> None:
        try:
            import fitz  # noqa: F401
        except ImportError:
            return

        for root, _, files in os.walk(self._folder):
            for fname in files:
                ext = os.path.splitext(fname)[1].lower()
                if ext not in _OFFICE_EXTS:
                    continue
                fpath = os.path.join(root, fname)
                cache = _office_cache_dir(fpath)
                # Skip if already cached
                if list(cache.glob("page_*.png")):
                    continue
                self._convert_one(fpath, cache, fitz)
        self.finished_all.emit()

    def _convert_one(self, file_path: str, cache_dir: Path, fitz) -> None:  # type: ignore[type-arg]
        try:
            result = subprocess.run(
                _soffice_cmd(self._soffice, cache_dir, file_path),
                capture_output=True,
                timeout=120,
            )
            if result.returncode != 0:
                return
            pdfs = list(cache_dir.glob("*.pdf"))
            if not pdfs:
                return
            with _silence_mupdf():
                doc = fitz.open(str(pdfs[0]))
            for i, page in enumerate(doc):
                mat = fitz.Matrix(2.0, 2.0)
                with _silence_mupdf():
                    pix = page.get_pixmap(matrix=mat, alpha=False)
                img = QImage(
                    pix.samples,
                    pix.width,
                    pix.height,
                    pix.stride,
                    QImage.Format.Format_RGB888,
                ).copy()
                img.save(str(cache_dir / f"page_{i:04d}.png"))
            doc.close()
        except Exception:
            pass


def preconvert_office_files(folder: str) -> _PreconvertWorker | None:
    """Start background pre-conversion of all Office files under *folder*.

    Returns the worker thread (caller must keep a reference to prevent GC),
    or ``None`` if LibreOffice is not installed.
    """
    soffice = _find_soffice()
    if soffice is None:
        return None
    worker = _PreconvertWorker(folder, soffice)
    worker.start()
    return worker
